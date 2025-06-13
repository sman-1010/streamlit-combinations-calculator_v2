import streamlit as st
import pandas as pd
from collections import Counter
from io import BytesIO

# Only import openpyxl if we do color-coding
# (We will conditionally import it later if needed)

def parse_list(input_str):
    """Helper to parse a comma-separated string of integers."""
    arr = []
    for x in input_str.split(','):
        x = x.strip()
        if x.lstrip('-').isdigit():  # handle negative if needed
            arr.append(int(x))
    return arr

def remove_nwis(sample_list, nwis):
    """Remove any items from sample_list that appear in nwis."""
    return [x for x in sample_list if x not in nwis]

def is_valid_triple(A, B, C, counts, strict_switch, nwis):
    """
    Check if the triple (A, B, C) can be formed with available counts.
    If strict_switch is True, apply extra checks involving nwis.
    """
    from collections import Counter
    temp_counts = Counter([A, B, C])
    
    # If strict_switch is on, apply extra NWIS logic
    if strict_switch:
        if (B - 5) in nwis:
            return False
        if (C - B + 5) in nwis:
            return False

    # Check we have enough occurrences for each number
    for num, req in temp_counts.items():
        if counts.get(num, 0) < req:
            return False
    return True

def is_valid_triple_single(M, S, T, Gen, Ext, counts, strict_switch, nwis):
    """
    Check if the triple (A, B, C) can be formed with available counts.
    If strict_switch is True, apply extra checks involving nwis.
    """
    from collections import Counter
    temp_counts = Counter([M, S, T, Gen, Ext])
    
    # If strict_switch is on, apply extra NWIS logic
    if strict_switch:
        if (M - 5) in nwis:
            return False

    # Check we have enough occurrences for each number
    for num, req in temp_counts.items():
        if counts.get(num, 0) < req:
            return False
    return True

def is_valid_triple_dual(M, S, T, Gen, Ext, counts, strict_switch, nwis):
    """
    Check if the triple (A, B, C) can be formed with available counts.
    If strict_switch is True, apply extra checks involving nwis.
    """
    from collections import Counter
    temp_counts = Counter([M, S, T, Gen, Ext])
    
    # If strict_switch is on, apply extra NWIS logic
    if strict_switch:
        if (M - 5) in nwis:
            return False
        if (S-M+5) in nwis:
            return False

    # Check we have enough occurrences for each number
    for num, req in temp_counts.items():
        if counts.get(num, 0) < req:
            return False
    return True

def is_valid_triple_double_dual(M, S, T, Gen, Ext, counts, strict_switch, nwis):
    """
    Check if the triple (A, B, C) can be formed with available counts.
    If strict_switch is True, apply extra checks involving nwis.
    """
    from collections import Counter
    temp_counts = Counter([M, S, T, Gen, Ext])
    
    # If strict_switch is on, apply extra NWIS logic
    if strict_switch:
        if (M - 15) in nwis:
            return False
        if (S - M + 15) in nwis:
            return False

    # Check we have enough occurrences for each number
    for num, req in temp_counts.items():
        if counts.get(num, 0) < req:
            return False
    return True

def is_valid_triple_double_single(M, S, T, Gen, Ext, counts, strict_switch, nwis):
    """
    Check if the triple (A, B, C) can be formed with available counts.
    If strict_switch is True, apply extra checks involving nwis.
    """
    from collections import Counter
    temp_counts = Counter([M, S, T, Gen, Ext])
    
    # If strict_switch is on, apply extra NWIS logic
    if strict_switch:
        if (S - 1) in nwis:
            return False

    # Check we have enough occurrences for each number
    for num, req in temp_counts.items():
        if counts.get(num, 0) < req:
            return False
    return True

def is_valid_double(A, B, counts, strict_switch, nwis):
    """
    Check if the double (A, B) can be formed with available counts.
    If strict_switch is True, apply extra checks involving nwis.
    """
    from collections import Counter
    temp_counts = Counter([A, B])
    
    # If strict_switch is on, apply extra NWIS logic
    if strict_switch:
        if (B - 1) in nwis:
            return False

    # Check we have enough occurrences for each number
    for num, req in temp_counts.items():
        if counts.get(num, 0) < req:
            return False
    return True

def compute_bins(triple, Main, G, R, C_list):
    """
    Calculate how many numbers in the triple come from
    each bin: [Main, G, R, C_list].
    """
    bins = [0, 0, 0, 0]  # [Priority_count, 2nd_count, 3rd_count, Backup_count]

    # Copies so original data is not modified
    main_copy = Main.copy()
    g_copy    = G.copy()
    r_copy    = R.copy()
    c_copy    = C_list.copy()

    for num in triple:
        if num in main_copy:
            bins[0] += 1
            main_copy.remove(num)
        elif num in g_copy:
            bins[1] += 1
            g_copy.remove(num)
        elif num in r_copy:
            bins[2] += 1
            r_copy.remove(num)
        elif num in c_copy:
            bins[3] += 1
            c_copy.remove(num)
    return bins

def main():
    st.title("Number Combinations Generator")

    # ---------------------------------------------------------------
    # 1) CONFIGURATION FLAGS - user selectable
    # ---------------------------------------------------------------

    # Alternative to segmented_control
    # Bold and large title with no spacing before radio buttons
    st.write("### **Select One of the Combinations Below:**")  # This avoids extra spacing

    # Radio button with hidden label to remove extra space
    method_selections = st.radio("Options:", ["Combination 1: Single", "Combination 2: Dual", "Combination 3: Double Single", "Combination 4: Double Dual"], label_visibility="collapsed")

    # st.write(f"You selected: {method_selection}")


    # method_selection = ''
    # if method_selections == "Combination 1: Triples":
    #     # st.write("You selected Combination 1: Triples")
    #     st.write("##### You selected Combination 1: Triples")
    #     st.write("""
    #     Three numbers (B,C,SUM) selected from Main,G,R,C with conditions
    #     - SUM = C + 5
    #     - Intermediate Values of B and C are greater than 0
    #     """)

    #     method_selection = 'triple'

    # elif method_selections == "Combination 2: Doubles":
    #     st.write("##### You selected Combination 2: Doubles")
    #     st.write("""
    #     Two numbers (B,SUM) selected from Main,G,R,C with conditions
    #     - SUM = B + 4
    #     - Intermediate Value of B is greater than 0
    #     """)
    #     method_selection = 'double'

    method_selection = ''
    if method_selections == "Combination 1: Single":
        method_selection = 'single'
        description = """
        <b>Three numbers (Main, Subsidary, Total) selected from Priority, 2nd, 3rd, Backup with conditions:</b>
        <ul>
            <li>Gen = X1 + 1</li>
            <li>Main = 5 + M1</li>
            <li>Subsidary = M1 + 1</li>
            <li>Exterior = Total - Main + 2</li>
            <li>Total = 5 + M1</li>
            <li>Intermediate Value of M1 is greater than 0</li>
        </ul>
        """
    elif method_selections == "Combination 2: Dual":
        method_selection = 'dual'
        description = """
        <b>Three numbers (Main, Subsidary, Total) selected from Priority, 2nd, 3rd, Backup with conditions:</b>
        <ul>
            <li>Gen = X1 + 1</li>
            <li>Main = 5 + M1</li>
            <li>Subsidary = M1 + M2</li>
            <li>Exterior = Total - Main + 1</li>
            <li>Total = 5 + M1 + M2</li>
            <li>Intermediate Values of M1 and M2 are greater than 0</li>
        </ul>
        """
    elif method_selections == "Combination 3: Double Single":
        method_selection = 'double_single'
        description = """
        <b>Three numbers (Main, Subsidary, Total) selected from Priority, 2nd, 3rd, Backup with conditions:</b>
        <ul>
            <li>Gen = X1 + 1</li>
            <li>Main = 8 + M1</li>
            <li>Subsidary = M1 + 1</li>
            <li>Exterior = Total - Main + 1</li>
            <li>Total = 4 + 8 + M1</li>
            <li>Intermediate Value of M1 is greater than 0</li>
        </ul>
        """
    elif method_selections == "Combination 4: Double Dual":
        method_selection = 'double_dual'
        description = """
        <b>Three numbers (Main, Subsidary, Total) selected from Priority, 2nd, 3rd, Backup with conditions:</b>
        <ul>
            <li>Gen = X1 + 1</li>
            <li>Main = 15 + M1</li>
            <li>Subsidary = M1 + M2</li>
            <li>Exterior = Total - Main</li>
            <li>Total = 15 + 15 + M1 + M2</li>
            <li>Intermediate Values of M1 and M2 are greater than 0</li>
        </ul>
        """

    # Detect Streamlit theme mode
    theme_mode = st.get_option("theme.base")  # 'light' or 'dark'

    # Set styles dynamically
    if theme_mode == "dark":
        bg_color = "#2b2b2b"  # Dark grey for dark mode
        text_color = "#ddd"  # Light text for dark mode
        border_color = "#444"  # Slightly lighter border
    else:
        bg_color = "#f0f0f0"  # Light grey for light mode
        text_color = "#333"  # Dark text for light mode
        border_color = "#d3d3d3"  # Light border

    # Styled HTML with dynamic theme colors
    st.markdown(
        f"""
        <div style="
            background-color: {bg_color}; 
            color: {text_color};
            padding: 15px; 
            border-radius: 10px; 
            border: 1px solid {border_color};
            box-shadow: 2px 2px 5px rgba(0,0,0,0.1);
            font-size: 16px;">
            <h4 style="color: {text_color};">You selected {method_selections}</h4>
            {description}
        </div>
        """,
        unsafe_allow_html=True
    )


    st.markdown("---")
    # st.subheader("Select Files to save below:")
    st.subheader("Not Wanted in List:")
    # save_valid = st.checkbox("Save valid combinations (valid_combinations_final.xlsx)", value=True)
    save_valid = True
    # save_no_color = st.checkbox("Save no-color version (visualized_no_color.xlsx)", value=True)
    save_no_color = False
    # save_with_color = st.checkbox("Save color version (visualized_with_color.xlsx)", value=True)
    save_with_color = True

    # Toggle buttons for user selection
    toggle_M_S = st.toggle("Strict Main and Subsidary", value=True)
    strict_switch = st.toggle("Strict Intermediate Values (M1 & M2)", value=False)
    toggle_T = st.toggle("Strict Total", value=False)    
    toggle_G = st.toggle("Strict Gen", value=False)    
    toggle_E = st.toggle("Strict Exterior", value=False)    

    st.markdown("---")
    st.subheader("Input Configuration")

    # ---------------------------------------------------------------
    # 2) DATA AND PARAMETERS - user editable
    # ---------------------------------------------------------------

    # strict_switch = st.checkbox("Enable strict_switch (NWIS check in intermediate steps)", value=False)
    # default_main = "1,3,5,9,11,13,15,16,21,23,24,25,29,31,32,33,35,37,39,41,45,47,48,52,57,65,67,68,82"
    default_main = "1,3,5,11,13,15,16,21,23,24,25,29,31,32,33,35,37,39,41,45,47,48,52,57,65,67,68,82"
    main_str = st.text_area("Priority list", default_main, height=80)

    # default_g = "3,5,6,11,13,15,16,24,31,32,35"
    default_g = "15,16,23,24,32,33,41"

    g_str = st.text_area("2nd list", default_g, height=80)

    # default_r = "4,10,12,14,15,16,20,22,23,24,28,32,33,34,41"
    default_r = "3,5,6,11,13,15,16,24,31,32,35"

    r_str = st.text_area("3rd list", default_r, height=80)

    # default_c_list = "12,14,22,32"
    default_c_list = "6,7,17,18,38,51,55,61,75"
    c_list_str = st.text_area("Backup list", default_c_list, height=80)

    # default_nwim = "2,4,9,10,12,14,19,20,26,27,28,34,36,42,43,44,46,49,50,53,54,56,58,59,60,62,64,66,69,70,72,73,74,76,78,79,80,21,23,26,28,29,33,39,22,4,10,12,22,28,34,4,9,10,14,19,20,28,34,44,9,10,14,19,20,22,28,30,34,40,44,50,53,54,56,59,60,70,80"
    default_nwim = "2,4,9,10,12,14,19,20,22,26,27,28,30,34,36,40,42,43,44,46,49,50,53,54,56,58,59,60,62,64,66,69,70,72,73,74,76,78,79,80"

    nwim_str = st.text_area("Not Wanted List", default_nwim, height=100)
    if st.button("Run Combinations Logic"):
        # Parse user inputs
        not_wanted_in_sum = parse_list(nwim_str)
        nwis = list(set(not_wanted_in_sum))

        Main   = parse_list(main_str)
        G      = parse_list(g_str)
        R      = parse_list(r_str)
        C_list = parse_list(c_list_str)

        # ---------------------------------------------------------------
        # 4) MAIN LOGIC
        # ---------------------------------------------------------------
        # 4a) Filter out NWIS items
        # Main   = remove_nwis(Main, nwis)
        # G      = remove_nwis(G, nwis)
        # R      = remove_nwis(R, nwis)
        # C_list = remove_nwis(C_list, nwis)

        # 4b) Create a major list and get counts
        major_list = Main + G + R + C_list
        counts = Counter(major_list)
        unique_sorted = sorted(counts.keys())
        
        if method_selection == 'single':

            valid_triples = []
            temp_skip = False
            Gen = 6
            Ext = 2
            if Gen in counts and Ext in counts:
                if toggle_G and Gen in nwis:
                    temp_skip = True                    
                if toggle_E and Ext in nwis:                
                    temp_skip = True                    

                if not temp_skip:
                    for M in unique_sorted:
                        if toggle_M_S and M in nwis:
                            continue

                        if not M > 5:
                            continue
                        
                        S = M - 4
                        if S not in counts:
                            continue
                        if toggle_M_S and S in nwis:
                            continue
                        
                        T = M
                        if toggle_T and T in nwis:
                            continue
                        if is_valid_triple_single(M, S, T, Gen, Ext, counts, strict_switch, nwis):
                            valid_triples.append((M, S, T, Gen, Ext))                            
            
            # 4d) Sort the triples and build a dataframe
            triple_bins = [
                (triple, compute_bins(triple, Main, G, R, C_list)) 
                for triple in valid_triples
            ]
            # triple_bins_sorted = sorted(triple_bins, key=lambda x: x[0][0])
            triple_bins_sorted = sorted(
                triple_bins,
                key=lambda x: (-x[1][0], -x[1][1], -x[1][2], -x[1][3], x[0])
            )            

            rows = []
            for triple, bins_count in triple_bins_sorted:
                # A, B, C_ = triple
                M, S, T, Gen, Ext = triple
                # Row: [B, C, A] + [Priority_count, 2nd_count, 3rd_count, Backup_count]
                rows.append([M, S, T, Gen, Ext] + bins_count)

            columns = ['Main', 'Subsidary', 'Total', 'Gen', 'Ext', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df = pd.DataFrame(rows, columns=columns)

            # Keep only rows that used exactly 3 items
            df = df[df[['Priority_count', '2nd_count', '3rd_count', 'Backup_count']].sum(axis=1) == 5]
            
            # With intermediate values
            rows = []
            for triple, bins_count in triple_bins_sorted:
                # M, S, T = triple
                M, S, T, Gen, Ext = triple

                # Row: [B, C, A] + [Priority_count, 2nd_count, 3rd_count, Backup_count]
                rows.append([M, S, T, Gen, Ext]+[M-5, 0] + bins_count)

            columns_triple = ['Main', 'Subsidary', 'Total', 'Gen', 'Ext','M1','M2', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df_with_inter = pd.DataFrame(rows, columns=columns_triple)

            # Keep only rows that used exactly 3 items
            df_with_inter = df_with_inter[df_with_inter[['Priority_count', '2nd_count', '3rd_count', 'Backup_count']].sum(axis=1) == 5]
            
            st.success("Combinations generated!")

            st.write(f"Number of valid rows: {len(df_with_inter)}")
            df_with_inter.index = range(1, len(df_with_inter) + 1)
            st.dataframe(df_with_inter)  # Show a sample

            # Creating the new DataFrame with empty spaces
            new_rows = []
            for _, row in df_with_inter.iterrows():
                new_rows.append([5, row['M1'], row['M2'],'', row['Total'], '', row['Priority_count'], row['2nd_count'], row['3rd_count'], row['Backup_count']])
                new_rows.append([row['Gen'], row['Main'], row['Subsidary'],row['Ext'], '', '', 'Main', '2nd', '3rd', 'Backup'])
                new_rows.append([''] * 10)  # Empty row

            # Creating the new DataFrame
            columns_triple = ['Gen', 'Main', 'Subsidary','Exterior','Total','--', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df_formatted = pd.DataFrame(new_rows, columns=columns_triple)
            df_formatted.index = range(1, len(df_formatted) + 1)
            st.write("Combinations Visualized:")
            st.dataframe(df_formatted)  # Show a sample

        elif method_selection == 'dual':

            valid_triples = []
            temp_skip = False
            Gen = 6
            if Gen in counts:
                if toggle_G and Gen in nwis:
                    temp_skip = True                    
                # if toggle_E and Ext in nwis:                
                #     temp_skip = True           
                
                if not temp_skip:
                    for M in unique_sorted:
                        if toggle_M_S and M in nwis:
                            continue
                        if not M > 5:
                            continue
                        
                        M1 = M - 5
                        for S in unique_sorted:
                            if not S > M1:
                                continue
                            
                            T = S + 5
                            if T not in counts:
                                continue                        
                            if toggle_T and T in nwis:
                                continue
                            
                            Ext = S - M + 6
                            if Ext not in counts:
                                continue          
                            if toggle_E and Ext in nwis:
                                continue
  
                            if is_valid_triple_dual(M, S, T, Gen, Ext, counts, strict_switch, nwis):
                                valid_triples.append((M, S, T, Gen, Ext))                                    

                         
            
            # 4d) Sort the triples and build a dataframe
            triple_bins = [
                (triple, compute_bins(triple, Main, G, R, C_list)) 
                for triple in valid_triples
            ]
            # Sort by bins in descending priority order
            triple_bins_sorted = sorted(
                triple_bins,
                key=lambda x: (-x[1][0], -x[1][1], -x[1][2], -x[1][3], x[0])
            )


            # triple_bins_sorted = sorted(triple_bins, key=lambda x: x[0][0])

            rows = []
            for triple, bins_count in triple_bins_sorted:
                # A, B, C_ = triple
                M, S, T, Gen, Ext = triple
                # Row: [B, C, A] + [Priority_count, 2nd_count, 3rd_count, Backup_count]
                rows.append([M, S, T, Gen, Ext] + bins_count)

            columns = ['Main', 'Subsidary', 'Total', 'Gen', 'Ext', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df = pd.DataFrame(rows, columns=columns)

            # Keep only rows that used exactly 3 items
            df = df[df[['Priority_count', '2nd_count', '3rd_count', 'Backup_count']].sum(axis=1) == 5]
            
            # With intermediate values
            rows = []
            for triple, bins_count in triple_bins_sorted:
                M, S, T, Gen, Ext = triple

                # Row: [B, C, A] + [Priority_count, 2nd_count, 3rd_count, Backup_count]
                rows.append([M, S, T, Gen, Ext]+[M-5, S-M+5] + bins_count)

            columns_triple = ['Main', 'Subsidary', 'Total', 'Gen', 'Ext','M1','M2', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df_with_inter = pd.DataFrame(rows, columns=columns_triple)

            # Keep only rows that used exactly 3 items
            df_with_inter = df_with_inter[df_with_inter[['Priority_count', '2nd_count', '3rd_count', 'Backup_count']].sum(axis=1) == 5]
            
            st.success("Combinations generated!")

            st.write(f"Number of valid rows: {len(df_with_inter)}")
            df_with_inter.index = range(1, len(df_with_inter) + 1)
            st.dataframe(df_with_inter)  # Show a sample

            # Creating the new DataFrame with empty spaces
            new_rows = []
            for _, row in df_with_inter.iterrows():
                # new_rows.append(['', row['Main'], row['Subsidary'], '', '', 'Main', '2nd', '3rd', 'Backup'])
                # new_rows.append([5, row['M1'], row['M2'], row['Total'], '', row['Priority_count'], row['2nd_count'], row['3rd_count'], row['Backup_count']])
                # new_rows.append([''] * 9)  # Empty row
                new_rows.append([5, row['M1'], row['M2'],'', row['Total'], '', row['Priority_count'], row['2nd_count'], row['3rd_count'], row['Backup_count']])
                new_rows.append([row['Gen'], row['Main'], row['Subsidary'],row['Ext'], '', '', 'Main', '2nd', '3rd', 'Backup'])
                new_rows.append([''] * 10)  # Empty row                

            # Creating the new DataFrame
            columns_triple = ['Gen', 'Main', 'Subsidary','Exterior','Total','--', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df_formatted = pd.DataFrame(new_rows, columns=columns_triple)
            df_formatted.index = range(1, len(df_formatted) + 1)
            st.write("Combinations Visualized:")
            st.dataframe(df_formatted)  # Show a sample
        
        elif method_selection == 'double_single':

            valid_triples = []
            temp_skip = False
            Gen = 6
            Ext = Gen
            if Gen in counts:
                if toggle_G and Gen in nwis:
                    temp_skip = True                    
                if toggle_E and Ext in nwis:                
                    temp_skip = True           
                
                if not temp_skip:            
                    for M in unique_sorted:
                        if toggle_M_S and M in nwis:
                            continue
                        if not M > 8:
                            continue
                        
                        S = M - 7
                        if S not in counts:
                            continue
                        if toggle_M_S and S in nwis:
                            continue

                        T = 5 + M
                        if T not in counts:
                            continue    
                        if toggle_T and T in nwis:
                            continue

                        # if is_valid_triple_dual(M, S, T, counts, strict_switch, nwis):
                        #     valid_triples.append((M, S, T))   
                        if is_valid_triple_double_single(M, S, T, Gen, Ext, counts, strict_switch, nwis):
                            valid_triples.append((M, S, T, Gen, Ext))                              

                         
            
            # 4d) Sort the triples and build a dataframe
            triple_bins = [
                (triple, compute_bins(triple, Main, G, R, C_list)) 
                for triple in valid_triples
            ]
            # Sort by bins in descending priority order
            triple_bins_sorted = sorted(
                triple_bins,
                key=lambda x: (-x[1][0], -x[1][1], -x[1][2], -x[1][3], x[0])
            )

            rows = []
            for triple, bins_count in triple_bins_sorted:
                # A, B, C_ = triple
                M, S, T, Gen, Ext = triple
                # Row: [B, C, A] + [Priority_count, 2nd_count, 3rd_count, Backup_count]
                rows.append([M, S, T, Gen, Ext] + bins_count)

            columns = ['Main', 'Subsidary', 'Total', 'Gen', 'Ext', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df = pd.DataFrame(rows, columns=columns)

            # Keep only rows that used exactly 3 items
            df = df[df[['Priority_count', '2nd_count', '3rd_count', 'Backup_count']].sum(axis=1) == 5]
            
            # With intermediate values
            rows = []
            for triple, bins_count in triple_bins_sorted:
                M, S, T, Gen, Ext = triple

                # Row: [B, C, A] + [Priority_count, 2nd_count, 3rd_count, Backup_count]
                rows.append([M, S, T, Gen, Ext]+[S-1] + bins_count)

            columns_triple = ['Main', 'Subsidary', 'Total', 'Gen', 'Ext', 'M1', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df_with_inter = pd.DataFrame(rows, columns=columns_triple)

            # Keep only rows that used exactly 3 items
            df_with_inter = df_with_inter[df_with_inter[['Priority_count', '2nd_count', '3rd_count', 'Backup_count']].sum(axis=1) == 5]
            
            st.success("Combinations generated!")

            st.write(f"Number of valid rows: {len(df_with_inter)}")
            df_with_inter.index = range(1, len(df_with_inter) + 1)
            st.dataframe(df_with_inter)  # Show a sample

            # Creating the new DataFrame with empty spaces
            new_rows = []
            for _, row in df_with_inter.iterrows():
                new_rows.append([5, 8, row['M1'],'', row['Total'], '', row['Priority_count'], row['2nd_count'], row['3rd_count'], row['Backup_count']])
                new_rows.append([row['Gen'], row['Main'], row['Subsidary'],row['Ext'], '', '', 'Main', '2nd', '3rd', 'Backup'])
                new_rows.append([''] * 10 )  # Empty row

            # Creating the new DataFrame
            columns_triple = ['Gen', 'Main', 'Subsidary','Exterior','Total','--', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df_formatted = pd.DataFrame(new_rows, columns=columns_triple)
            df_formatted.index = range(1, len(df_formatted) + 1)
            st.write("Combinations Visualized:")
            st.dataframe(df_formatted)  # Show a sample


        elif method_selection == 'double_dual':

            valid_triples = []
            temp_skip = False
            Gen = 6
            if Gen in counts:
                if toggle_G and Gen in nwis:
                    temp_skip = True                    
                # if toggle_E and Ext in nwis:                
                #     temp_skip = True           
                
                if not temp_skip:            
                    for M in unique_sorted:
                        if toggle_M_S and M in nwis:
                            continue
                        if not M > 15:
                            continue
                        
                        M1 = M - 15
                        for S in unique_sorted:
                            if not S > M1:
                                continue
                            
                            T = S + 20
                            if T not in counts:
                                continue                        
                            if toggle_T and T in nwis:
                                continue

                            Ext = S - M + 20
                            if Ext not in counts:
                                continue          
                            if toggle_E and Ext in nwis:
                                continue

                            if is_valid_triple_double_dual(M, S, T, Gen, Ext, counts, strict_switch, nwis):
                                valid_triples.append((M, S, T, Gen, Ext))   

                         
            
            # 4d) Sort the triples and build a dataframe
            triple_bins = [
                (triple, compute_bins(triple, Main, G, R, C_list)) 
                for triple in valid_triples
            ]
            # Sort by bins in descending priority order
            triple_bins_sorted = sorted(
                triple_bins,
                key=lambda x: (-x[1][0], -x[1][1], -x[1][2], -x[1][3], x[0])
            )

            rows = []
            for triple, bins_count in triple_bins_sorted:
                # A, B, C_ = triple
                M, S, T, Gen, Ext = triple
                # Row: [B, C, A] + [Priority_count, 2nd_count, 3rd_count, Backup_count]
                rows.append([M, S, T, Gen, Ext] + bins_count)

            columns = ['Main', 'Subsidary', 'Total', 'Gen', 'Ext', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df = pd.DataFrame(rows, columns=columns)

            # Keep only rows that used exactly 3 items
            df = df[df[['Priority_count', '2nd_count', '3rd_count', 'Backup_count']].sum(axis=1) == 5]
            
            # With intermediate values
            rows = []
            for triple, bins_count in triple_bins_sorted:
                M, S, T, Gen, Ext = triple

                # Row: [B, C, A] + [Priority_count, 2nd_count, 3rd_count, Backup_count]
                rows.append([M, S, T, Gen, Ext]+[M-15, S-M+15] + bins_count)

            columns_triple = ['Main', 'Subsidary', 'Total', 'Gen', 'Ext','M1','M2', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df_with_inter = pd.DataFrame(rows, columns=columns_triple)

            # Keep only rows that used exactly 3 items
            df_with_inter = df_with_inter[df_with_inter[['Priority_count', '2nd_count', '3rd_count', 'Backup_count']].sum(axis=1) == 5]
            
            st.success("Combinations generated!")

            st.write(f"Number of valid rows: {len(df_with_inter)}")
            df_with_inter.index = range(1, len(df_with_inter) + 1)
            st.dataframe(df_with_inter)  # Show a sample

            # Creating the new DataFrame with empty spaces
            new_rows = []
            for _, row in df_with_inter.iterrows():
                new_rows.append([5,15, row['M1'], row['M2'],'', row['Total'], '', row['Priority_count'], row['2nd_count'], row['3rd_count'], row['Backup_count']])
                new_rows.append([6, row['Main'], row['Subsidary'], '',row['Ext'],'', '', 'Main', '2nd', '3rd', 'Backup'])
                new_rows.append([''] * 10)  # Empty row

            # Creating the new DataFrame
            columns_triple = ['Gen', 'Main', 'Subsidary','---','Exterior','Total','--', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df_formatted = pd.DataFrame(new_rows, columns=columns_triple)
            df_formatted.index = range(1, len(df_formatted) + 1)
            st.write("Combinations Visualized:")
            st.dataframe(df_formatted)  # Show a sample
                    
        else:
            # 4c) Generate valid triples
            valid_doubles = []
            for B in unique_sorted:
                if toggle_M_S and B in nwis:
                    continue                
                A = B + 4  # Condition from the original code
                if A not in counts:
                    continue
                if toggle_T and A in nwis:
                    continue                
                if B > 1:
                    if is_valid_double(A, B, counts, strict_switch, nwis):
                        valid_doubles.append((A, B))                        

        
            double_bins = [
                (double, compute_bins(double, Main, G, R, C_list)) 
                for double in valid_doubles
            ]
            double_bins_sorted = sorted(double_bins, key=lambda x: x[0][0])

            rows = []
            for double, bins_count in double_bins_sorted:
                A, B = double
                # Row: [B, C, A] + [Priority_count, 2nd_count, 3rd_count, Backup_count]
                rows.append([B, '', A] + bins_count)

            columns = ['B', '', 'SUM', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df = pd.DataFrame(rows, columns=columns)
            # Keep only rows that used exactly 3 items
            df = df[df[['Priority_count', '2nd_count', '3rd_count', 'Backup_count']].sum(axis=1) == 2]
            
            rows = []
            for double, bins_count in double_bins_sorted:
                A, B = double
                # Row: [B, C, A] + [Priority_count, 2nd_count, 3rd_count, Backup_count]
                rows.append([B, '', A,B-1] + bins_count)

            columns = ['B', '', 'SUM','M1', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df_with_inter = pd.DataFrame(rows, columns=columns)
            # Keep only rows that used exactly 3 items
            df_with_inter = df_with_inter[df_with_inter[['Priority_count', '2nd_count', '3rd_count', 'Backup_count']].sum(axis=1) == 2]
            

            st.success("Combinations generated Double!")
            st.write(f"Number of valid rows: {len(df_with_inter)}")
            df_with_inter.index = range(1, len(df_with_inter) + 1)
            st.dataframe(df_with_inter)  # Show a sample

            # Creating the new DataFrame with empty spaces
            new_rows = []
            for _, row in df_with_inter.iterrows():
                new_rows.append(['', row['B'], '', '', '', 'Main', '2nd', '3rd', 'Backup'])
                new_rows.append([5, row['M1'], '', row['SUM'], '', row['Priority_count'], row['2nd_count'], row['3rd_count'], row['Backup_count']])
                new_rows.append([''] * 9)  # Empty row

            # Creating the new DataFrame
            columns_triple = ['-', 'B', '','SUM','--', 'Priority_count', '2nd_count', '3rd_count', 'Backup_count']
            df_formatted = pd.DataFrame(new_rows, columns=columns_triple)
            df_formatted.index = range(1, len(df_formatted) + 1)
            st.write("Combinations Visualized:")
            st.dataframe(df_formatted)  # Show a sample


            # ---------------------------------------------------------------
            # 5) SAVE FIRST FILE: valid_combinations_final.xlsx (in memory)
            # ---------------------------------------------------------------
        if save_valid:
            valid_buffer = BytesIO()
            # Adjust file name for strict if needed
            valid_filename = "valid_combinations.xlsx"
            if strict_switch:
                valid_filename = "valid_combinations_strict.xlsx"      
            valid_filename = f"{method_selection}_{valid_filename}"      
            df_with_inter.to_excel(valid_buffer, index=False)
            valid_buffer.seek(0)
            st.download_button(
                label=f"Download {valid_filename}",
                data=valid_buffer,
                file_name=valid_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # ---------------------------------------------------------------
        # 7) SAVE THIRD FILE: visualized_with_color.xlsx (in memory)
        # ---------------------------------------------------------------
        if save_with_color:
            # We'll build the colored Excel directly with openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill
            from openpyxl.styles import Alignment


            wb = Workbook()
            ws = wb.active
            center = Alignment(horizontal="center", vertical="center")

            # Define color fills
            green_fill  = PatternFill(start_color="92D051", fill_type="solid")
            blue_fill   = PatternFill(start_color="06B0F0", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", fill_type="solid")
            light_blue_fill = PatternFill(start_color="CAEDFB", fill_type="solid")
            purple_fill = PatternFill(start_color="D86DCD", fill_type="solid")
            tea_green_fill = PatternFill(start_color="C0F0C8", fill_type="solid")
            orange_fill = PatternFill(start_color="FFBF00", fill_type="solid")
            peach_fill = PatternFill(start_color="F1A983", fill_type="solid")

            # We want to replicate the same 3-row block structure
            idx = 0  # to iterate over df rows in groups of 1 (each row = one triple in df)
            row_idx = 1
            if method_selection == 'single':
                while idx < len(df):
                    M_val = df.iloc[idx]['Main']
                    S_val = df.iloc[idx]['Subsidary']
                    T_val = df.iloc[idx]['Total']
                    G_val = df.iloc[idx]['Gen']
                    E_val = df.iloc[idx]['Ext']
                    M1_val = df_with_inter.iloc[idx]['M1']

                    ws.cell(row=row_idx, column=1, value="")
                    ws.cell(row=row_idx, column=2, value="X").fill = blue_fill
                    ws.cell(row=row_idx, column=3, value="M1").fill = yellow_fill
                    ws.cell(row=row_idx, column=4, value="")
                    ws.cell(row=row_idx, column=5, value="")
                    ws.cell(row=row_idx, column=6, value="Dual").fill = purple_fill
                    for col in range(1, 12):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center
                    
                    row_idx = row_idx + 1
                    # The top row
                    ws.cell(row=row_idx, column=1, value="(M,S,T,G,E)")
                    ws.cell(row=row_idx, column=2, value=5).fill = blue_fill
                    ws.cell(row=row_idx, column=3, value=M1_val).fill = yellow_fill
                    ws.cell(row=row_idx, column=4, value="")
                    ws.cell(row=row_idx, column=5, value="")
                    ws.cell(row=row_idx, column=6, value=T_val).fill = orange_fill
                    ws.cell(row=row_idx, column=7, value="")
                    ws.cell(row=row_idx, column=8, value="Priority")
                    ws.cell(row=row_idx, column=9, value="2nd")
                    ws.cell(row=row_idx, column=10, value="3rd")
                    ws.cell(row=row_idx, column=11, value="Backup")
                    for col in range(1, 12):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center
                    
                    # The second row
                    triple_str = f"({M_val}, {S_val}, {T_val}, {G_val}, {E_val})"
                    row_idx = row_idx + 1
                    ws.cell(row=row_idx, column=1, value=triple_str)
                    ws.cell(row=row_idx, column=2, value=G_val).fill = light_blue_fill
                    ws.cell(row=row_idx, column=3, value=M_val).fill = purple_fill
                    ws.cell(row=row_idx, column=4, value=S_val).fill = green_fill
                    ws.cell(row=row_idx, column=5, value=E_val).fill = tea_green_fill
                    ws.cell(row=row_idx, column=6, value="").fill = orange_fill
                    ws.cell(row=row_idx, column=7, value="")
                    # ws.cell(row=row_2, column=5, value=T_val).fill = yellow_fill

                    # We also want to show counts in columns 7..10:
                    ws.cell(row=row_idx, column=8, value=df.iloc[idx]['Priority_count'])
                    ws.cell(row=row_idx, column=9, value=df.iloc[idx]['2nd_count'])
                    ws.cell(row=row_idx, column=10, value=df.iloc[idx]['3rd_count'])
                    ws.cell(row=row_idx, column=11, value=df.iloc[idx]['Backup_count'])
                    for col in range(1, 12):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center

                    row_idx+=1
                    ws.cell(row=row_idx, column=1, value="")
                    ws.cell(row=row_idx, column=2, value="Gen").fill = light_blue_fill
                    ws.cell(row=row_idx, column=3, value="Main").fill = purple_fill
                    ws.cell(row=row_idx, column=4, value="Subsidary").fill = green_fill
                    ws.cell(row=row_idx, column=5, value="Exterior").fill = tea_green_fill
                    ws.cell(row=row_idx, column=6, value="Total").fill = orange_fill
                    for col in range(1, 12):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center
                    # The third row is blank
                    for c in range(1, 12):
                        ws.cell(row=row_idx+2, column=c, value="")

                    # Move to the next triple
                    row_idx += 3
                    idx += 1

            elif method_selection == 'dual':
                while idx < len(df):
                    M_val = df.iloc[idx]['Main']
                    S_val = df.iloc[idx]['Subsidary']
                    T_val = df.iloc[idx]['Total']
                    G_val = df.iloc[idx]['Gen']
                    E_val = df.iloc[idx]['Ext']
                    M1_val = df_with_inter.iloc[idx]['M1']
                    M2_val = df_with_inter.iloc[idx]['M2']                    

                    ws.cell(row=row_idx, column=1, value="")
                    ws.cell(row=row_idx, column=2, value="X").fill = blue_fill
                    ws.cell(row=row_idx, column=3, value="M1").fill = yellow_fill
                    ws.cell(row=row_idx, column=4, value="M2").fill = peach_fill
                    ws.cell(row=row_idx, column=5, value="")
                    ws.cell(row=row_idx, column=6, value="Single").fill = purple_fill
                    for col in range(1, 12):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center
                    
                    row_idx = row_idx + 1
                    # The top row
                    ws.cell(row=row_idx, column=1, value="(M,S,T,G,E)")
                    ws.cell(row=row_idx, column=2, value=5).fill = blue_fill
                    ws.cell(row=row_idx, column=3, value=M1_val).fill = yellow_fill
                    ws.cell(row=row_idx, column=4, value=M2_val).fill = peach_fill
                    ws.cell(row=row_idx, column=5, value="")
                    ws.cell(row=row_idx, column=6, value=T_val).fill = orange_fill
                    ws.cell(row=row_idx, column=7, value="")
                    ws.cell(row=row_idx, column=8, value="Priority")
                    ws.cell(row=row_idx, column=9, value="2nd")
                    ws.cell(row=row_idx, column=10, value="3rd")
                    ws.cell(row=row_idx, column=11, value="Backup")
                    for col in range(1, 12):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center
                    
                    # The second row
                    triple_str = f"({M_val}, {S_val}, {T_val}, {G_val}, {E_val})"
                    row_idx = row_idx + 1
                    ws.cell(row=row_idx, column=1, value=triple_str)
                    ws.cell(row=row_idx, column=2, value=G_val).fill = light_blue_fill
                    ws.cell(row=row_idx, column=3, value=M_val).fill = purple_fill
                    ws.cell(row=row_idx, column=4, value=S_val).fill = green_fill
                    ws.cell(row=row_idx, column=5, value=E_val).fill = tea_green_fill
                    ws.cell(row=row_idx, column=6, value="").fill = orange_fill
                    ws.cell(row=row_idx, column=7, value="")
                    # ws.cell(row=row_2, column=5, value=T_val).fill = yellow_fill

                    # We also want to show counts in columns 7..10:
                    ws.cell(row=row_idx, column=8, value=df.iloc[idx]['Priority_count'])
                    ws.cell(row=row_idx, column=9, value=df.iloc[idx]['2nd_count'])
                    ws.cell(row=row_idx, column=10, value=df.iloc[idx]['3rd_count'])
                    ws.cell(row=row_idx, column=11, value=df.iloc[idx]['Backup_count'])
                    for col in range(1, 12):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center

                    row_idx+=1
                    ws.cell(row=row_idx, column=1, value="")
                    ws.cell(row=row_idx, column=2, value="Gen").fill = light_blue_fill
                    ws.cell(row=row_idx, column=3, value="Main").fill = purple_fill
                    ws.cell(row=row_idx, column=4, value="Subsidary").fill = green_fill
                    ws.cell(row=row_idx, column=5, value="Exterior").fill = tea_green_fill
                    ws.cell(row=row_idx, column=6, value="Total").fill = orange_fill
                    for col in range(1, 12):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center
                    # The third row is blank
                    for c in range(1, 12):
                        ws.cell(row=row_idx+2, column=c, value="")

                    # Move to the next triple
                    row_idx += 3
                    idx += 1
            elif method_selection == 'double_single':
                while idx < len(df):
                    M_val = df.iloc[idx]['Main']
                    S_val = df.iloc[idx]['Subsidary']
                    T_val = df.iloc[idx]['Total']
                    G_val = df.iloc[idx]['Gen']
                    E_val = df.iloc[idx]['Ext']
                    M1_val = df_with_inter.iloc[idx]['M1']

                    ws.cell(row=row_idx, column=1, value="")
                    ws.cell(row=row_idx, column=2, value="X1").fill = blue_fill
                    ws.cell(row=row_idx, column=3, value="X2").fill = blue_fill
                    ws.cell(row=row_idx, column=4, value="M1").fill = yellow_fill
                    ws.cell(row=row_idx, column=5, value="")
                    ws.cell(row=row_idx, column=6, value="Double Single").fill = purple_fill
                    for col in range(1, 12):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center
                    
                    row_idx = row_idx + 1
                    # The top row
                    ws.cell(row=row_idx, column=1, value="(M,S,T,G,E)")
                    ws.cell(row=row_idx, column=2, value=5).fill = blue_fill
                    ws.cell(row=row_idx, column=3, value=8).fill = blue_fill
                    ws.cell(row=row_idx, column=4, value=M1_val).fill = yellow_fill
                    ws.cell(row=row_idx, column=5, value="")
                    ws.cell(row=row_idx, column=6, value=T_val).fill = orange_fill
                    ws.cell(row=row_idx, column=7, value="")
                    ws.cell(row=row_idx, column=8, value="Priority")
                    ws.cell(row=row_idx, column=9, value="2nd")
                    ws.cell(row=row_idx, column=10, value="3rd")
                    ws.cell(row=row_idx, column=11, value="Backup")
                    for col in range(1, 12):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center
                    
                    # The second row
                    triple_str = f"({M_val}, {S_val}, {T_val}, {G_val}, {E_val})"
                    row_idx = row_idx + 1
                    ws.cell(row=row_idx, column=1, value=triple_str)
                    ws.cell(row=row_idx, column=2, value=G_val).fill = light_blue_fill
                    ws.cell(row=row_idx, column=3, value=M_val).fill = purple_fill
                    ws.cell(row=row_idx, column=4, value=S_val).fill = green_fill
                    ws.cell(row=row_idx, column=5, value=E_val).fill = tea_green_fill
                    ws.cell(row=row_idx, column=6, value="").fill = orange_fill
                    ws.cell(row=row_idx, column=7, value="")
                    # ws.cell(row=row_2, column=5, value=T_val).fill = yellow_fill

                    # We also want to show counts in columns 7..10:
                    ws.cell(row=row_idx, column=8, value=df.iloc[idx]['Priority_count'])
                    ws.cell(row=row_idx, column=9, value=df.iloc[idx]['2nd_count'])
                    ws.cell(row=row_idx, column=10, value=df.iloc[idx]['3rd_count'])
                    ws.cell(row=row_idx, column=11, value=df.iloc[idx]['Backup_count'])
                    for col in range(1, 12):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center

                    row_idx+=1
                    ws.cell(row=row_idx, column=1, value="")
                    ws.cell(row=row_idx, column=2, value="Gen").fill = light_blue_fill
                    ws.cell(row=row_idx, column=3, value="Main").fill = purple_fill
                    ws.cell(row=row_idx, column=4, value="Subsidary").fill = green_fill
                    ws.cell(row=row_idx, column=5, value="Exterior").fill = tea_green_fill
                    ws.cell(row=row_idx, column=6, value="Total").fill = orange_fill
                    for col in range(1, 12):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center
                    # The third row is blank
                    for c in range(1, 12):
                        ws.cell(row=row_idx+2, column=c, value="")

                    # Move to the next triple
                    row_idx += 3
                    idx += 1

            elif method_selection == 'double_dual':
                while idx < len(df):
                    M_val = df.iloc[idx]['Main']
                    S_val = df.iloc[idx]['Subsidary']
                    T_val = df.iloc[idx]['Total']
                    G_val = df.iloc[idx]['Gen']
                    E_val = df.iloc[idx]['Ext']
                    M1_val = df_with_inter.iloc[idx]['M1']
                    M2_val = df_with_inter.iloc[idx]['M2']

                    ws.cell(row=row_idx, column=1, value="")
                    ws.cell(row=row_idx, column=2, value="X1").fill = blue_fill
                    ws.cell(row=row_idx, column=3, value="X2").fill = blue_fill
                    ws.cell(row=row_idx, column=4, value="M1").fill = yellow_fill
                    ws.cell(row=row_idx, column=5, value="M2").fill = peach_fill
                    ws.cell(row=row_idx, column=6, value="")
                    ws.cell(row=row_idx, column=7, value="Double Dual").fill = purple_fill
                    for col in range(1, 13):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center
                    
                    row_idx = row_idx + 1
                    # The top row
                    ws.cell(row=row_idx, column=1, value="(M,S,T,G,E)")
                    ws.cell(row=row_idx, column=2, value=5).fill = blue_fill
                    ws.cell(row=row_idx, column=3, value=15).fill = blue_fill
                    ws.cell(row=row_idx, column=4, value=M1_val).fill = yellow_fill
                    ws.cell(row=row_idx, column=5, value=M2_val).fill = peach_fill
                    ws.cell(row=row_idx, column=6, value="")
                    ws.cell(row=row_idx, column=7, value=T_val).fill = orange_fill
                    ws.cell(row=row_idx, column=8, value="")
                    ws.cell(row=row_idx, column=9, value="Priority")
                    ws.cell(row=row_idx, column=10, value="2nd")
                    ws.cell(row=row_idx, column=11, value="3rd")
                    ws.cell(row=row_idx, column=12, value="Backup")
                    for col in range(1, 13):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center
                    
                    # The second row
                    triple_str = f"({M_val}, {S_val}, {T_val}, {G_val}, {E_val})"
                    row_idx = row_idx + 1
                    ws.cell(row=row_idx, column=1, value=triple_str)
                    ws.cell(row=row_idx, column=2, value=G_val).fill = light_blue_fill
                    ws.cell(row=row_idx, column=3, value=M_val).fill = purple_fill
                    ws.cell(row=row_idx, column=4, value=S_val).fill = green_fill
                    ws.cell(row=row_idx, column=5, value="").fill = green_fill
                    ws.cell(row=row_idx, column=6, value=E_val).fill = tea_green_fill
                    ws.cell(row=row_idx, column=7, value="").fill = orange_fill
                    ws.cell(row=row_idx, column=8, value="")
                    ws.cell(row=row_idx, column=9, value=df.iloc[idx]['Priority_count'])
                    ws.cell(row=row_idx, column=10, value=df.iloc[idx]['2nd_count'])
                    ws.cell(row=row_idx, column=11, value=df.iloc[idx]['3rd_count'])
                    ws.cell(row=row_idx, column=12, value=df.iloc[idx]['Backup_count'])
                    for col in range(1, 13):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center

                    row_idx+=1
                    ws.cell(row=row_idx, column=1, value="")
                    ws.cell(row=row_idx, column=2, value="Gen").fill = light_blue_fill
                    ws.cell(row=row_idx, column=3, value="Main").fill = purple_fill
                    ws.cell(row=row_idx, column=4, value="Subsidary").fill = green_fill
                    ws.cell(row=row_idx, column=5, value="").fill = green_fill
                    ws.cell(row=row_idx, column=6, value="Exterior").fill = tea_green_fill
                    ws.cell(row=row_idx, column=7, value="Total").fill = orange_fill
                    for col in range(1, 13):               # 1–11 in your snippet
                        ws.cell(row=row_idx, column=col).alignment = center
                    # The third row is blank
                    for c in range(1, 13):
                        ws.cell(row=row_idx+2, column=c, value="")

                    # Move to the next triple
                    row_idx += 3
                    idx += 1
            else:
                while idx < len(df):
                    B_val = df.iloc[idx]['B']
                    SUM_val = df.iloc[idx]['SUM']

                    # The top row
                    ws.cell(row=row_idx, column=1, value="")
                    ws.cell(row=row_idx, column=2, value="")
                    ws.cell(row=row_idx, column=3, value=B_val).fill = green_fill
                    ws.cell(row=row_idx, column=4, value="")
                    ws.cell(row=row_idx, column=5, value="")
                    ws.cell(row=row_idx, column=6, value="")
                    ws.cell(row=row_idx, column=7, value="Main")
                    ws.cell(row=row_idx, column=8, value="G")
                    ws.cell(row=row_idx, column=9, value="R")
                    ws.cell(row=row_idx, column=10, value="C")

                    # The second row
                    triple_str = f"({B_val}, {SUM_val})"
                    row_2 = row_idx + 1
                    ws.cell(row=row_2, column=1, value=triple_str)
                    ws.cell(row=row_2, column=2, value=5)
                    ws.cell(row=row_2, column=3, value=(B_val - 1))
                    # ws.cell(row=row_2, column=4, value=(C_val - (B_val - 5)))
                    ws.cell(row=row_2, column=4, value="")
                    ws.cell(row=row_2, column=5, value=SUM_val).fill = yellow_fill

                    # We also want to show counts in columns 7..10:
                    ws.cell(row=row_2, column=7, value=df.iloc[idx]['Priority_count'])
                    ws.cell(row=row_2, column=8, value=df.iloc[idx]['2nd_count'])
                    ws.cell(row=row_2, column=9, value=df.iloc[idx]['3rd_count'])
                    ws.cell(row=row_2, column=10, value=df.iloc[idx]['Backup_count'])

                    # The third row is blank
                    for c in range(1, 11):
                        ws.cell(row=row_idx+2, column=c, value="")

                    # Move to the next triple
                    row_idx += 3
                    idx += 1

            color_buffer = BytesIO()
            color_filename = "visualized_with_color.xlsx"
            if strict_switch:
                color_filename = "visualized_with_color_strict.xlsx"
            color_filename = f"{method_selection}_{color_filename}"
            wb.save(color_buffer)
            color_buffer.seek(0)

            st.download_button(
                label=f"Download {color_filename}",
                data=color_buffer,
                file_name=color_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # st.success("All done!")

if __name__ == "__main__":
    main()
